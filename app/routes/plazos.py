"""Plazos de pago — CRUD completo + exportar/importar Excel."""
import io
from datetime import datetime

from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, send_file)
from flask_login import login_required

from app.models import db, Plazo
from app.utils.decorators import admin_required, editor_required

plazos_bp = Blueprint('plazos', __name__)


# ─── LISTADO ──────────────────────────────────────────────────────────────────

@plazos_bp.route('/plazos')
@login_required
def plazos():
    all_plazos = Plazo.query.order_by(Plazo.days).all()
    return render_template('plazos.html', plazos=all_plazos)


# ─── CREAR ────────────────────────────────────────────────────────────────────

@plazos_bp.route('/plazos/nuevo', methods=['POST'])
@login_required
@editor_required
def new_plazo():
    name = request.form.get('name', '').strip()
    if not name:
        flash('El nombre del plazo es obligatorio.', 'error')
        return redirect(url_for('plazos.plazos'))
    if Plazo.query.filter_by(name=name).first():
        flash(f'Ya existe un plazo con el nombre "{name}".', 'error')
        return redirect(url_for('plazos.plazos'))
    try:
        days = int(request.form.get('days', 0) or 0)
        if days < 0:
            raise ValueError
    except (ValueError, TypeError):
        flash('Los días deben ser un número entero positivo.', 'error')
        return redirect(url_for('plazos.plazos'))

    db.session.add(Plazo(
        name=name,
        days=days,
        description=request.form.get('description', '').strip(),
    ))
    db.session.commit()
    flash(f'Plazo "{name}" creado correctamente.', 'success')
    return redirect(url_for('plazos.plazos'))


# ─── EDITAR ───────────────────────────────────────────────────────────────────

@plazos_bp.route('/plazos/<int:id>/editar', methods=['POST'])
@login_required
@editor_required
def edit_plazo(id):
    p = Plazo.query.get_or_404(id)
    name = request.form.get('name', '').strip()
    if not name:
        flash('El nombre del plazo es obligatorio.', 'error')
        return redirect(url_for('plazos.plazos'))
    # Duplicate name check (excluding self)
    dup = Plazo.query.filter(Plazo.name == name, Plazo.id != id).first()
    if dup:
        flash(f'Ya existe un plazo con el nombre "{name}".', 'error')
        return redirect(url_for('plazos.plazos'))
    try:
        days = int(request.form.get('days', 0) or 0)
        if days < 0:
            raise ValueError
    except (ValueError, TypeError):
        flash('Los días deben ser un número entero positivo.', 'error')
        return redirect(url_for('plazos.plazos'))

    p.name        = name
    p.days        = days
    p.description = request.form.get('description', '').strip()
    db.session.commit()
    flash(f'Plazo "{p.name}" actualizado.', 'success')
    return redirect(url_for('plazos.plazos'))


# ─── ELIMINAR ─────────────────────────────────────────────────────────────────

@plazos_bp.route('/plazos/<int:id>/eliminar', methods=['POST'])
@login_required
@admin_required
def delete_plazo(id):
    p = Plazo.query.get_or_404(id)
    name = p.name
    db.session.delete(p)
    db.session.commit()
    flash(f'Plazo "{name}" eliminado.', 'success')
    return redirect(url_for('plazos.plazos'))


# ─── EXPORTAR EXCEL ───────────────────────────────────────────────────────────

@plazos_bp.route('/plazos/exportar')
@login_required
def export_plazos_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('plazos.plazos'))

    all_plazos = Plazo.query.order_by(Plazo.days).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Plazos de Pago'

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='7C3AED')
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left     = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin     = Border(
        left=Side(style='thin', color='DDD6FE'),  right=Side(style='thin', color='DDD6FE'),
        top=Side(style='thin', color='DDD6FE'),   bottom=Side(style='thin', color='DDD6FE'),
    )
    alt_fill   = PatternFill('solid', fgColor='F5F3FF')
    white_fill = PatternFill('solid', fgColor='FFFFFF')

    headers    = ['ID', 'Nombre', 'Días', 'Tipo', 'Descripción']
    col_widths = [8,    28,        10,     14,      40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    for i, p in enumerate(all_plazos, 2):
        fill = alt_fill if i % 2 == 0 else white_fill
        tipo = 'Contado' if p.days == 0 else 'Crédito'
        row  = [p.id, p.name, p.days, tipo, p.description or '']
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border    = thin
            cell.fill      = fill
            cell.alignment = center if col_idx in [1, 3] else left

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d')
    return send_file(buf, download_name=f'plazos_pago_{fecha}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── DESCARGAR PLANTILLA ──────────────────────────────────────────────────────

@plazos_bp.route('/plazos/plantilla')
@login_required
def export_plazos_template():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('plazos.plazos'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Plazos de Pago'

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='7C3AED')
    center   = Alignment(horizontal='center', vertical='center')
    thin     = Border(
        left=Side(style='thin', color='DDD6FE'),  right=Side(style='thin', color='DDD6FE'),
        top=Side(style='thin', color='DDD6FE'),   bottom=Side(style='thin', color='DDD6FE'),
    )

    headers    = ['Nombre *', 'Días', 'Descripción']
    col_widths = [28,          10,     40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 24

    # Example rows
    examples = [
        ('Contado', 0, 'Pago en el momento de la entrega'),
        ('30 días', 30, 'Factura a 30 días de la fecha de emisión'),
        ('60 días', 60, ''),
        ('90 días', 90, ''),
    ]
    for r, (name, days, desc) in enumerate(examples, 2):
        for col_idx, val in enumerate([name, days, desc], 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = Font(color='94A3B8', italic=True)

    # Instructions sheet
    ws2 = wb.create_sheet('Instrucciones')
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 50
    instr = [
        ('INSTRUCCIONES', ''),
        ('', ''),
        ('Nombre *', 'Obligatorio. Nombre único del plazo (ej: "30 días").'),
        ('Días', 'Número de días. Usar 0 para "Contado". Por defecto: 0.'),
        ('Descripción', 'Texto opcional con condiciones adicionales.'),
        ('', ''),
        ('Nota:', 'Si ya existe un plazo con el mismo nombre, se actualizará.'),
    ]
    for r, (k, v) in enumerate(instr, 1):
        ca = ws2.cell(row=r, column=1, value=k)
        ws2.cell(row=r, column=2, value=v)
        if r == 1:
            ca.font = Font(bold=True, size=12, color='7C3AED')
        elif v and r > 2:
            ca.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='plantilla_plazos.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── IMPORTAR EXCEL ───────────────────────────────────────────────────────────

@plazos_bp.route('/plazos/importar', methods=['POST'])
@login_required
@editor_required
def import_plazos_excel():
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('plazos.plazos'))

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Por favor subí un archivo Excel (.xlsx).', 'error')
        return redirect(url_for('plazos.plazos'))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().rstrip(' *') if cell.value else ''
                   for cell in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        if 'Nombre' not in col:
            flash('No se encontró la columna "Nombre". Usá la plantilla oficial.', 'error')
            return redirect(url_for('plazos.plazos'))

        created = updated = skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            def get(col_name, default=''):
                idx = col.get(col_name)
                if idx is None:
                    return default
                val = row[idx]
                return val if val is not None else default

            name = str(get('Nombre', '')).strip()
            if not name:
                skipped += 1
                continue

            try:
                days = int(float(get('Días', 0) or 0))
                if days < 0:
                    days = 0
            except (ValueError, TypeError):
                days = 0

            desc = str(get('Descripción', '')).strip()

            existing = Plazo.query.filter_by(name=name).first()
            if existing:
                existing.days        = days
                existing.description = desc
                updated += 1
            else:
                db.session.add(Plazo(name=name, days=days, description=desc))
                created += 1

        db.session.commit()
        msg = f'Importación completada: {created} creados, {updated} actualizados'
        if skipped:
            msg += f', {skipped} omitidos'
        flash(msg, 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al procesar el archivo: {e}', 'error')

    return redirect(url_for('plazos.plazos'))
