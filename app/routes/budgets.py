"""Budget routes."""
import io
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user

from app.models import db, Budget, BudgetItem, Product, Category
from app.utils.pdf import build_budget_pdf
from app.utils.email import notify_budget_created, notify_whatsapp_budget, send_budget_to_client

budgets_bp = Blueprint('budgets', __name__)


@budgets_bp.route('/presupuestos')
@login_required
def budgets():
    page = request.args.get('page', 1, type=int)
    bgs = Budget.query.order_by(Budget.created_at.desc()).paginate(page=page, per_page=20)
    return render_template('budgets.html', budgets=bgs)


@budgets_bp.route('/presupuestos/nuevo', methods=['GET', 'POST'])
@login_required
def new_budget():
    if request.method == 'POST':
        data = request.get_json()
        if not data or not data.get('client_name', '').strip():
            return jsonify({'error': 'El nombre del cliente es obligatorio.'}), 400
        try:
            discount = float(data.get('discount', 0))
            tax = float(data.get('tax', 21))
            if not 0 <= discount <= 100 or not 0 <= tax <= 100:
                raise ValueError
        except (ValueError, TypeError):
            return jsonify({'error': 'Descuento o IVA inválidos.'}), 400

        budget = Budget(
            client_name=data['client_name'].strip(),
            client_email=data.get('client_email', ''),
            client_phone=data.get('client_phone', ''),
            client_address=data.get('client_address', ''),
            notes=data.get('notes', ''),
            discount=discount,
            tax=tax,
            user_id=current_user.id
        )
        budget.generate_number()
        db.session.add(budget)
        db.session.flush()

        for item_data in data.get('items', []):
            try:
                qty = float(item_data['quantity'])
                price = float(item_data['unit_price'])
                if qty <= 0 or price < 0:
                    raise ValueError
            except (ValueError, TypeError, KeyError):
                db.session.rollback()
                return jsonify({'error': 'Ítem con valores inválidos.'}), 400
            db.session.add(BudgetItem(
                budget_id=budget.id,
                product_id=item_data.get('product_id') or None,
                description=item_data['description'],
                quantity=qty,
                unit_price=price
            ))

        db.session.commit()
        notify_budget_created(budget)
        notify_whatsapp_budget(budget)
        return jsonify({'id': budget.id, 'number': budget.budget_number})

    prods = Product.query.filter_by(is_active=True).order_by(Product.name).all()
    categories = Category.query.all()
    # Serialize to dict so tojson works in the template
    products_data = [
        {'id': p.id, 'name': p.name, 'price': p.price, 'stock': p.stock,
         'unit': p.unit, 'sku': p.sku or '', 'category_id': p.category_id or 0}
        for p in prods
    ]
    return render_template('budget_form.html', products=products_data, categories=categories)


@budgets_bp.route('/presupuestos/<int:id>')
@login_required
def view_budget(id):
    budget = Budget.query.get_or_404(id)
    return render_template('budget_view.html', budget=budget)


@budgets_bp.route('/presupuestos/<int:id>/pdf')
@login_required
def budget_pdf(id):
    budget = Budget.query.get_or_404(id)
    pdf_bytes = build_budget_pdf(budget)
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        download_name=f"presupuesto_{budget.budget_number}.pdf"
    )


@budgets_bp.route('/presupuestos/<int:id>/estado', methods=['POST'])
@login_required
def update_budget_status(id):
    budget = Budget.query.get_or_404(id)
    new_status = request.form.get('status', '')
    valid_statuses = {'borrador', 'enviado', 'aprobado', 'rechazado'}
    if new_status not in valid_statuses:
        flash('Estado inválido.', 'error')
        return redirect(url_for('budgets.view_budget', id=id))
    budget.status = new_status
    db.session.commit()
    flash('Estado actualizado.', 'success')
    return redirect(url_for('budgets.view_budget', id=id))


@budgets_bp.route('/presupuestos/<int:id>/enviar-email', methods=['POST'])
@login_required
def send_budget_email(id):
    budget = Budget.query.get_or_404(id)
    if not budget.client_email:
        flash('El cliente no tiene email registrado.', 'error')
        return redirect(url_for('budgets.view_budget', id=id))
    # Single call to the shared PDF builder — no duplication
    pdf_bytes = build_budget_pdf(budget)
    ok = send_budget_to_client(budget, pdf_bytes)
    if ok:
        budget.status = 'enviado'
        db.session.commit()
        flash(f'Presupuesto enviado por email a {budget.client_email}.', 'success')
    else:
        flash('No se pudo enviar el email. Verificá la configuración de correo.', 'error')
    return redirect(url_for('budgets.view_budget', id=id))


@budgets_bp.route('/presupuestos/exportar')
@login_required
def export_budgets_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('budgets.budgets'))

    bgs = Budget.query.order_by(Budget.created_at.desc()).all()
    wb = openpyxl.Workbook()

    # ── Sheet 1: Budgets ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Presupuestos'
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='7C3AED')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_al = Alignment(horizontal='right', vertical='center')
    thin = Border(
        left=Side(style='thin', color='DDD6FE'), right=Side(style='thin', color='DDD6FE'),
        top=Side(style='thin', color='DDD6FE'), bottom=Side(style='thin', color='DDD6FE'),
    )
    alt_fill = PatternFill('solid', fgColor='F5F3FF')
    white_f = PatternFill('solid', fgColor='FFFFFF')
    STATUS_LABELS = {'borrador': 'Borrador', 'enviado': 'Enviado',
                     'aprobado': 'Aprobado', 'rechazado': 'Rechazado'}

    headers = ['N° Presupuesto', 'Fecha', 'Cliente', 'Email', 'Teléfono',
               'Dirección', 'Subtotal', 'Descuento %', 'IVA %', 'Total', 'Estado', 'Notas']
    col_widths = [18, 16, 28, 28, 16, 32, 14, 13, 10, 14, 12, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    for i, b in enumerate(bgs, 2):
        fill = alt_fill if i % 2 == 0 else white_f
        row_data = [b.budget_number, b.created_at.strftime('%d/%m/%Y'), b.client_name,
                    b.client_email or '', b.client_phone or '', b.client_address or '',
                    b.subtotal, b.discount, b.tax, b.total,
                    STATUS_LABELS.get(b.status, b.status), b.notes or '']
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin
            cell.fill = fill
            if col_idx in [7, 10]:
                cell.number_format = '$#,##0.00'
                cell.alignment = right_al
            elif col_idx in [8, 9]:
                cell.number_format = '0.00"%"'
                cell.alignment = center
            elif col_idx in [1, 2, 11]:
                cell.alignment = center
            else:
                cell.alignment = left

    last = len(bgs) + 2
    ws.cell(row=last, column=9, value='TOTAL GENERAL').font = Font(bold=True)
    tc = ws.cell(row=last, column=10, value=sum(b.total for b in bgs))
    tc.font = Font(bold=True, color='7C3AED')
    tc.number_format = '$#,##0.00'
    tc.alignment = right_al

    # ── Sheet 2: Clients ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Clientes')
    hdr2_fill = PatternFill('solid', fgColor='0F766E')
    hdrs2 = ['Cliente', 'Email', 'Teléfono', 'Dirección', 'Cant. Presupuestos', 'Total Facturado']
    widths2 = [30, 30, 16, 36, 20, 18]
    for col, (h, w) in enumerate(zip(hdrs2, widths2), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr2_fill
        cell.alignment = center
        cell.border = thin
        ws2.column_dimensions[cell.column_letter].width = w
    ws2.row_dimensions[1].height = 24
    ws2.freeze_panes = 'A2'

    from collections import defaultdict
    clients = defaultdict(lambda: {'email': '', 'phone': '', 'address': '', 'count': 0, 'total': 0.0})
    for b in bgs:
        key = (b.client_name, b.client_email or '')
        clients[key]['email'] = b.client_email or ''
        clients[key]['phone'] = b.client_phone or ''
        clients[key]['address'] = b.client_address or ''
        clients[key]['count'] += 1
        clients[key]['total'] += b.total

    for i, ((name, _), data) in enumerate(sorted(clients.items()), 2):
        fill = alt_fill if i % 2 == 0 else white_f
        row_data2 = [name, data['email'], data['phone'], data['address'], data['count'], data['total']]
        for col_idx, val in enumerate(row_data2, 1):
            cell = ws2.cell(row=i, column=col_idx, value=val)
            cell.border = thin
            cell.fill = fill
            if col_idx == 6:
                cell.number_format = '$#,##0.00'
                cell.alignment = right_al
            elif col_idx == 5:
                cell.alignment = center
            else:
                cell.alignment = left

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d')
    return send_file(buf, download_name=f'presupuestos_{fecha}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
