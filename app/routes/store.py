"""Public store and store admin routes."""
import json
from datetime import datetime
from threading import Thread
import os

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required

from app.models import db, Product, Category, ProductStore, StoreConfig, Order, OrderItem
from app.utils.decorators import admin_required, editor_required
from app.utils.email import send_email, email_template, send_whatsapp

store_bp = Blueprint('store', __name__)


def get_store_config():
    cfg = StoreConfig.query.first()
    if not cfg:
        cfg = StoreConfig()
        db.session.add(cfg)
        db.session.commit()
    return cfg


# ─── PUBLIC STORE ─────────────────────────────────────────────────────────────

@store_bp.route('/tienda')
def store():
    cfg = get_store_config()
    q = request.args.get('q', '')
    cat_id = request.args.get('cat', '')
    min_price = request.args.get('min_price', '')
    max_price = request.args.get('max_price', '')
    sort = request.args.get('sort', 'nombre')
    featured_only = request.args.get('featured', '')

    query = db.session.query(Product, ProductStore).join(
        ProductStore, Product.id == ProductStore.product_id
    ).filter(Product.is_active == True, ProductStore.visible == True, Product.stock > 0)

    if q:
        query = query.filter(Product.name.ilike(f'%{q}%'))
    if cat_id:
        query = query.filter(Product.category_id == cat_id)
    if min_price:
        query = query.filter(Product.price >= float(min_price))
    if max_price:
        query = query.filter(Product.price <= float(max_price))
    if featured_only:
        query = query.filter(ProductStore.featured == True)
    if sort == 'precio_asc':
        query = query.order_by(Product.price.asc())
    elif sort == 'precio_desc':
        query = query.order_by(Product.price.desc())
    else:
        query = query.order_by(ProductStore.sort_order.asc(), Product.name.asc())

    results = query.all()
    categories = Category.query.all()
    featured = db.session.query(Product, ProductStore).join(
        ProductStore, Product.id == ProductStore.product_id
    ).filter(Product.is_active == True, ProductStore.visible == True,
             ProductStore.featured == True, Product.stock > 0).all()

    return render_template('store/store.html',
        cfg=cfg, products=results, categories=categories,
        featured=featured, q=q, cat_id=cat_id,
        min_price=min_price, max_price=max_price, sort=sort)


@store_bp.route('/tienda/producto/<int:id>')
def store_product(id):
    cfg = get_store_config()
    product = Product.query.get_or_404(id)
    store_info = ProductStore.query.filter_by(product_id=id).first()
    if not store_info or not store_info.visible:
        return redirect(url_for('store.store'))
    related = db.session.query(Product, ProductStore).join(
        ProductStore, Product.id == ProductStore.product_id
    ).filter(Product.category_id == product.category_id, Product.id != id,
             Product.is_active == True, ProductStore.visible == True, Product.stock > 0
    ).limit(4).all()
    return render_template('store/product.html', cfg=cfg, product=product,
                           store_info=store_info, related=related)


@store_bp.route('/tienda/pedido', methods=['POST'])
def store_order():
    cfg = get_store_config()
    data = request.get_json()
    if not data or not data.get('client_name', '').strip():
        return jsonify({'error': 'Nombre del cliente obligatorio'}), 400

    order = Order(
        client_name=data['client_name'].strip(),
        client_email=data.get('client_email', ''),
        client_phone=data.get('client_phone', ''),
        client_address=data.get('client_address', ''),
        notes=data.get('notes', ''),
        total=float(data.get('total', 0))
    )
    order.generate_number()
    db.session.add(order)
    db.session.flush()
    for item in data.get('items', []):
        db.session.add(OrderItem(
            order_id=order.id,
            product_id=item.get('product_id'),
            description=item['description'],
            quantity=int(item['quantity']),
            unit_price=float(item['unit_price']),
            color=item.get('color', ''),
            size=item.get('size', '')
        ))
    db.session.commit()

    # WhatsApp to admin
    wa_phone = os.environ.get('WHATSAPP_PHONE', '')
    if wa_phone:
        msg = (
            f"🛒 *Nuevo Pedido {order.order_number}*\n\n"
            f"👤 *Cliente:* {order.client_name}\n"
            f"📱 *Tel:* {order.client_phone}\n"
            f"💵 *Total:* ${order.total:,.2f}\n"
            f"📦 *Items:* {len(order.items)}"
        )
        Thread(target=send_whatsapp, args=(wa_phone, msg)).start()

    # Email to admins
    from app.models import User
    from app import flask_app
    if flask_app.config.get('MAIL_ENABLED'):
        admins = [u.email for u in User.query.filter_by(role='admin', is_active_user=True).all() if u.email]
        if admins:
            items_html = ''.join([
                f"<tr><td style='padding:8px 12px;border-bottom:1px solid #e2e8f0'>{i.description}</td>"
                f"<td style='padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:center'>{i.quantity}</td>"
                f"<td style='padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:right'>${i.subtotal:,.2f}</td></tr>"
                for i in order.items
            ])
            content = f"""
            <div style="background:#d1fae5;border:1px solid #6ee7b7;border-radius:8px;padding:20px;margin-bottom:20px">
              <div style="font-size:28px;margin-bottom:8px">🛒</div>
              <div style="font-size:16px;font-weight:600;color:#065f46">Nuevo pedido desde la tienda</div>
            </div>
            <table style="width:100%;border-collapse:collapse;margin-bottom:16px;font-size:14px">
              <tr><td style="padding:8px 12px;color:#64748b;font-weight:600">Cliente</td><td>{order.client_name}</td></tr>
              <tr style="background:#f8fafc"><td style="padding:8px 12px;color:#64748b;font-weight:600">Teléfono</td><td>{order.client_phone}</td></tr>
              <tr><td style="padding:8px 12px;color:#64748b;font-weight:600">Email</td><td>{order.client_email or '-'}</td></tr>
            </table>
            <table style="width:100%;border-collapse:collapse;font-size:14px">
              <thead><tr style="background:#f97316"><th style="padding:10px 12px;color:white;text-align:left">Producto</th><th style="padding:10px 12px;color:white;text-align:center">Cant.</th><th style="padding:10px 12px;color:white;text-align:right">Subtotal</th></tr></thead>
              <tbody>{items_html}</tbody>
              <tfoot><tr><td colspan="2" style="padding:10px 12px;font-weight:700;text-align:right;color:#f97316">TOTAL:</td><td style="padding:10px 12px;font-weight:700;color:#f97316;text-align:right">${order.total:,.2f}</td></tr></tfoot>
            </table>
            """
            send_email(
                subject=f"🛒 Nuevo pedido {order.order_number} — {order.client_name}",
                recipients=admins,
                html_body=email_template(f"Pedido {order.order_number}", content, '#f97316')
            )

    wa_number = cfg.whatsapp_number.replace('+', '').replace(' ', '') if cfg.whatsapp_number else ''
    wa_items = '%0A'.join([f"- {i.description} x{i.quantity}" for i in order.items])
    wa_msg = (
        f"Hola! Quiero hacer el siguiente pedido (N° {order.order_number}):%0A%0A"
        f"{wa_items}%0A%0A*Total: ${order.total:,.2f}*%0A%0A"
        f"Nombre: {order.client_name}%0ATel: {order.client_phone}"
    )
    wa_link = f"https://wa.me/{wa_number}?text={wa_msg}" if wa_number else ''
    return jsonify({'order_number': order.order_number, 'wa_link': wa_link})


@store_bp.route('/tienda/mp/crear-preferencia', methods=['POST'])
def mp_create_preference():
    cfg = get_store_config()
    if not cfg.mp_enabled or not cfg.mp_access_token:
        return jsonify({'error': 'MercadoPago no configurado'}), 400
    data = request.get_json()
    import urllib.request as urllib_req
    payload = {
        "items": [{"title": f"Pedido {data.get('order_number')} — {data.get('buyer_name')}",
                   "quantity": 1, "unit_price": float(data.get('total', 0)), "currency_id": "ARS"}],
        "payer": {"name": data.get('buyer_name'), "email": data.get('buyer_email') or "cliente@tienda.com"},
        "external_reference": data.get('order_number', ''),
        "back_urls": {
            "success": cfg.mp_success_url or request.host_url + "tienda/mp/gracias",
            "failure": request.host_url + "tienda",
            "pending": request.host_url + "tienda"
        },
        "auto_return": "approved",
        "statement_descriptor": cfg.store_name[:22],
    }
    try:
        req = urllib_req.Request(
            "https://api.mercadopago.com/checkout/preferences",
            data=json.dumps(payload).encode(),
            headers={"Content-Type": "application/json",
                     "Authorization": f"Bearer {cfg.mp_access_token}"}
        )
        resp = urllib_req.urlopen(req, timeout=15)
        result = json.loads(resp.read())
        return jsonify({"init_point": result.get("init_point"), "id": result.get("id")})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@store_bp.route('/tienda/mp/gracias')
def mp_success():
    cfg = get_store_config()
    order_number = request.args.get('external_reference', '')
    return render_template('store/mp_success.html', cfg=cfg, order_number=order_number)


# ─── STORE ADMIN ──────────────────────────────────────────────────────────────

@store_bp.route('/admin/tienda')
@login_required
@admin_required
def store_admin():
    cfg = get_store_config()
    orders = Order.query.order_by(Order.created_at.desc()).limit(20).all()
    prods = Product.query.filter_by(is_active=True).all()
    store_products = [(p, ProductStore.query.filter_by(product_id=p.id).first()) for p in prods]
    return render_template('store/admin.html', cfg=cfg, orders=orders, store_products=store_products)


@store_bp.route('/admin/tienda/config', methods=['POST'])
@login_required
@admin_required
def store_config_save():
    cfg = get_store_config()
    fields = [
        'store_name', 'store_slogan', 'store_description', 'whatsapp_number',
        'primary_color', 'secondary_color', 'logo_url', 'banner_url',
        'contact_email', 'address', 'instagram', 'facebook',
        'banner_title', 'banner_subtitle', 'featured_title',
        'about_title', 'about_text', 'about_image_url',
        'announcement_text', 'announcement_color', 'footer_text',
        'mp_public_key', 'mp_access_token', 'mp_success_url',
        'transfer_cbu', 'transfer_alias', 'transfer_bank', 'transfer_owner',
    ]
    for field in fields:
        val = request.form.get(field)
        if val is not None:
            setattr(cfg, field, val)
    # Booleans
    cfg.show_stock = 'show_stock' in request.form
    cfg.show_sku = 'show_sku' in request.form
    cfg.about_enabled = 'about_enabled' in request.form
    cfg.announcement_enabled = 'announcement_enabled' in request.form
    cfg.mp_enabled = 'mp_enabled' in request.form
    cfg.transfer_enabled = 'transfer_enabled' in request.form
    db.session.commit()
    flash('Configuración de la tienda guardada.', 'success')
    return redirect(url_for('store.store_admin'))


@store_bp.route('/admin/tienda/producto/<int:id>', methods=['POST'])
@login_required
@editor_required
def store_product_save(id):
    si = ProductStore.query.filter_by(product_id=id).first()
    if not si:
        si = ProductStore(product_id=id)
        db.session.add(si)
    si.visible = 'visible' in request.form
    si.featured = 'featured' in request.form
    si.original_price = float(request.form.get('original_price') or 0)
    si.image_url = request.form.get('image_url', '')
    si.image_url_2 = request.form.get('image_url_2', '')
    si.image_url_3 = request.form.get('image_url_3', '')
    si.badge = request.form.get('badge', '')
    si.colors = request.form.get('colors', '')
    si.sizes = request.form.get('sizes', '')
    si.store_description = request.form.get('store_description', '')
    si.sort_order = int(request.form.get('sort_order') or 0)
    db.session.commit()
    flash('Producto actualizado en la tienda.', 'success')
    return redirect(url_for('store.store_admin'))


@store_bp.route('/admin/tienda/pedido/<int:id>/estado', methods=['POST'])
@login_required
def update_order_status(id):
    order = Order.query.get_or_404(id)
    old_status = order.status
    new_status = request.form.get('status', '')
    valid = {'nuevo', 'en proceso', 'confirmado', 'enviado', 'cancelado'}
    if new_status not in valid:
        flash('Estado inválido.', 'error')
        return redirect(url_for('store.store_orders'))
    order.status = new_status
    db.session.commit()
    if order.client_phone and order.status != old_status:
        cfg = get_store_config()
        emoji = {'nuevo': '🆕', 'en proceso': '⚙️', 'confirmado': '✅', 'enviado': '🚚', 'cancelado': '❌'}.get(order.status, '📦')
        msg = (
            f"{emoji} *{cfg.store_name} — Tu pedido fue actualizado*\n\n"
            f"📋 *Pedido:* {order.order_number}\n"
            f"📦 *Estado:* {order.status.upper()}\n\n"
            f"Gracias por tu compra, {order.client_name}!"
        )
        Thread(target=send_whatsapp, args=(order.client_phone, msg)).start()
    flash('Estado del pedido actualizado.', 'success')
    return redirect(url_for('store.store_orders'))


@store_bp.route('/admin/tienda/pedidos')
@login_required
def store_orders():
    orders = Order.query.order_by(Order.created_at.desc()).all()
    return render_template('store/orders.html', orders=orders)


@store_bp.route('/admin/tienda/pedido/<int:id>')
@login_required
def order_detail(id):
    order = Order.query.get_or_404(id)
    return render_template('store/order_detail.html', order=order)


@store_bp.route('/admin/tienda/pedidos/exportar')
@login_required
@admin_required
def export_orders_excel():
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('store.store_orders'))

    orders = Order.query.order_by(Order.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pedidos'
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='F97316')
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'),
    )

    headers = ['N° Pedido', 'Fecha', 'Cliente', 'Teléfono', 'Email',
               'Dirección', 'Productos', 'Total', 'Estado', 'Notas']
    col_widths = [15, 18, 22, 16, 28, 30, 40, 12, 14, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for row_idx, o in enumerate(orders, 2):
        items_str = ' | '.join([f"{i.description} x{i.quantity}" for i in o.items])
        fill = PatternFill('solid', fgColor='FFF7ED') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        row_data = [o.order_number, o.created_at.strftime('%d/%m/%Y %H:%M'), o.client_name,
                    o.client_phone or '', o.client_email or '', o.client_address or '',
                    items_str, o.total, o.status, o.notes or '']
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin
            cell.fill = fill
            if col_idx == 8:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_idx in [1, 9]:
                cell.alignment = center

    last = len(orders) + 2
    ws.cell(row=last, column=7, value='TOTAL GENERAL').font = Font(bold=True)
    tc = ws.cell(row=last, column=8, value=sum(o.total for o in orders))
    tc.font = Font(bold=True, color='F97316')
    tc.number_format = '$#,##0.00'
    tc.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='pedidos_kairos.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
