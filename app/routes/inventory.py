"""Inventory routes: products, categories, stock movements, Excel import/export."""
import io
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user

from app.models import db, Product, Category, StockMovement, utcnow
from app.utils.decorators import admin_required, editor_required
from app.utils.email import notify_low_stock_alert, notify_whatsapp_low_stock

inventory_bp = Blueprint('inventory', __name__)


# ─── PRODUCTS ─────────────────────────────────────────────────────────────────

@inventory_bp.route('/productos')
@login_required
def products():
    q = request.args.get('q', '')
    cat_id = request.args.get('cat', '')
    query = Product.query.filter_by(is_active=True)
    if q:
        query = query.filter(Product.name.ilike(f'%{q}%'))
    if cat_id:
        query = query.filter_by(category_id=cat_id)
    prods = query.order_by(Product.name).all()
    categories = Category.query.all()
    return render_template('products.html', products=prods, categories=categories, q=q, cat_id=cat_id)


@inventory_bp.route('/productos/nuevo', methods=['GET', 'POST'])
@login_required
@editor_required
def new_product():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('El nombre del producto es obligatorio.', 'error')
            return redirect(url_for('inventory.new_product'))
        try:
            price = float(request.form.get('price') or 0)
            cost = float(request.form.get('cost') or 0)
            stock = int(request.form.get('stock') or 0)
            min_stock = int(request.form.get('min_stock') or 5)
        except (ValueError, TypeError):
            flash('Los valores numéricos son inválidos.', 'error')
            return redirect(url_for('inventory.new_product'))

        product = Product(
            name=name,
            sku=request.form.get('sku') or None,
            description=request.form.get('description', ''),
            category_id=request.form.get('category_id') or None,
            price=price, cost=cost, stock=stock,
            min_stock=min_stock,
            unit=request.form.get('unit', 'unidad')
        )
        db.session.add(product)
        db.session.flush()
        if product.stock > 0:
            db.session.add(StockMovement(
                product_id=product.id, user_id=current_user.id,
                movement_type='entrada', quantity=product.stock,
                previous_stock=0, new_stock=product.stock, notes='Stock inicial'
            ))
        db.session.commit()
        flash('Producto creado exitosamente.', 'success')
        return redirect(url_for('inventory.products'))
    categories = Category.query.all()
    return render_template('product_form.html', product=None, categories=categories)


@inventory_bp.route('/productos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@editor_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('El nombre del producto es obligatorio.', 'error')
            return redirect(url_for('inventory.edit_product', id=id))
        try:
            product.price = float(request.form.get('price') or 0)
            product.cost = float(request.form.get('cost') or 0)
            product.min_stock = int(request.form.get('min_stock') or 5)
        except (ValueError, TypeError):
            flash('Los valores numéricos son inválidos.', 'error')
            return redirect(url_for('inventory.edit_product', id=id))
        product.name = name
        product.sku = request.form.get('sku') or None
        product.description = request.form.get('description', '')
        product.category_id = request.form.get('category_id') or None
        product.unit = request.form.get('unit', 'unidad')
        product.updated_at = utcnow()
        db.session.commit()
        flash('Producto actualizado.', 'success')
        return redirect(url_for('inventory.products'))
    categories = Category.query.all()
    return render_template('product_form.html', product=product, categories=categories)


@inventory_bp.route('/productos/<int:id>/stock', methods=['POST'])
@login_required
@editor_required
def update_stock(id):
    product = Product.query.get_or_404(id)
    data = request.get_json()
    movement_type = data.get('type')
    try:
        quantity = int(data.get('quantity', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Cantidad inválida'}), 400
    if quantity <= 0:
        return jsonify({'error': 'La cantidad debe ser mayor a cero'}), 400
    notes = data.get('notes', '')
    prev = product.stock

    if movement_type == 'entrada':
        product.stock += quantity
    elif movement_type == 'salida':
        if product.stock < quantity:
            return jsonify({'error': 'Stock insuficiente'}), 400
        product.stock -= quantity
    elif movement_type == 'ajuste':
        product.stock = quantity
    else:
        return jsonify({'error': 'Tipo de movimiento inválido'}), 400

    db.session.add(StockMovement(
        product_id=product.id, user_id=current_user.id,
        movement_type=movement_type, quantity=quantity,
        previous_stock=prev, new_stock=product.stock, notes=notes
    ))
    product.updated_at = utcnow()
    db.session.commit()

    if product.low_stock and not (prev <= product.min_stock):
        notify_low_stock_alert(product)
        notify_whatsapp_low_stock(product)
    return jsonify({'stock': product.stock, 'low_stock': product.low_stock})


@inventory_bp.route('/productos/<int:id>/eliminar', methods=['POST'])
@login_required
@admin_required
def delete_product(id):
    product = Product.query.get_or_404(id)
    product.is_active = False
    db.session.commit()
    flash('Producto eliminado.', 'success')
    return redirect(url_for('inventory.products'))


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

@inventory_bp.route('/productos/exportar')
@login_required
def export_products_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('inventory.products'))

    prods = Product.query.filter_by(is_active=True).order_by(Product.name).all()
    categories = {c.id: c.name for c in Category.query.all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2563EB')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    thin = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'),
    )
    alt_fill = PatternFill('solid', fgColor='EFF6FF')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    low_fill = PatternFill('solid', fgColor='FEF2F2')

    headers = ['ID', 'Nombre', 'SKU', 'Categoría', 'Descripción',
               'Precio Venta', 'Costo', 'Stock', 'Stock Mínimo', 'Unidad', 'Activo']
    col_widths = [8, 30, 15, 20, 35, 14, 14, 10, 14, 12, 8]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    for i, p in enumerate(prods, 2):
        fill = low_fill if p.low_stock else (alt_fill if i % 2 == 0 else white_fill)
        row_data = [p.id, p.name, p.sku or '', categories.get(p.category_id, ''),
                    p.description or '', p.price, p.cost, p.stock, p.min_stock, p.unit, 'Sí']
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin
            cell.fill = fill
            if col_idx in [6, 7]:
                cell.number_format = '$#,##0.00'
                cell.alignment = right
            elif col_idx in [1, 8, 9, 11]:
                cell.alignment = center
            else:
                cell.alignment = left

    # Instructions sheet
    ws2 = wb.create_sheet('Instrucciones_Importar')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 55
    instr = [
        ('INSTRUCCIONES DE IMPORTACIÓN', ''),
        ('', ''),
        ('Columnas requeridas:', ''),
        ('Nombre', 'Obligatorio.'),
        ('SKU', 'Opcional. Si ya existe, se actualiza el producto.'),
        ('Categoría', 'Opcional. Debe coincidir con una categoría existente.'),
        ('Descripción', 'Opcional.'),
        ('Precio Venta', 'Número. Ej: 1500.50'),
        ('Costo', 'Número. Ej: 800'),
        ('Stock', 'Número entero. Solo para productos nuevos.'),
        ('Stock Mínimo', 'Número entero. Por defecto: 5'),
        ('Unidad', 'Ej: unidad, kg, litro, caja'),
        ('', ''),
        ('Nota:', 'El ID es ignorado al importar.'),
    ]
    for r, (k, v) in enumerate(instr, 1):
        ca = ws2.cell(row=r, column=1, value=k)
        ws2.cell(row=r, column=2, value=v)
        if r == 1:
            ca.font = Font(bold=True, size=12, color='2563EB')
        elif k and not v and r > 2:
            ca.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d')
    return send_file(buf, download_name=f'productos_kairos_{fecha}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── EXCEL IMPORT ─────────────────────────────────────────────────────────────

@inventory_bp.route('/productos/importar', methods=['POST'])
@login_required
@editor_required
def import_products_excel():
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('inventory.products'))

    file = request.files.get('excel_file')
    if not file or not file.filename.endswith(('.xlsx', '.xlsm')):
        flash('Por favor subí un archivo Excel (.xlsx).', 'error')
        return redirect(url_for('inventory.products'))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        if 'Nombre' not in col:
            flash('Columna "Nombre" no encontrada. Verificá el archivo.', 'error')
            return redirect(url_for('inventory.products'))

        categories_map = {c.name.lower(): c.id for c in Category.query.all()}
        created = updated = skipped = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue

            def get(col_name, default=''):
                idx = col.get(col_name)
                return (row[idx] if row[idx] is not None else default) if idx is not None else default

            name = str(get('Nombre', '')).strip()
            if not name:
                continue

            sku = str(get('SKU', '')).strip() or None
            cat_name = str(get('Categoría', '')).strip().lower()
            cat_id = categories_map.get(cat_name) if cat_name else None

            try:
                price = float(get('Precio Venta', 0) or 0)
                cost = float(get('Costo', 0) or 0)
                stock = int(float(get('Stock', 0) or 0))
                min_stk = int(float(get('Stock Mínimo', 5) or 5))
                if price < 0 or cost < 0 or stock < 0 or min_stk < 0:
                    raise ValueError("Valores negativos no permitidos")
            except (ValueError, TypeError) as e:
                errors.append(f'Fila {row_num}: {e}')
                skipped += 1
                continue

            unit = str(get('Unidad', 'unidad')).strip() or 'unidad'
            desc = str(get('Descripción', '')).strip()

            existing = None
            if sku:
                existing = Product.query.filter_by(sku=sku).first()
            if not existing:
                existing = Product.query.filter(
                    Product.name.ilike(name), Product.is_active == True
                ).first()

            if existing:
                existing.name = name
                existing.sku = sku
                existing.description = desc
                existing.category_id = cat_id
                existing.price = price
                existing.cost = cost
                existing.min_stock = min_stk
                existing.unit = unit
                existing.updated_at = utcnow()
                updated += 1
            else:
                p = Product(name=name, sku=sku, description=desc,
                            category_id=cat_id, price=price, cost=cost,
                            stock=stock, min_stock=min_stk, unit=unit)
                db.session.add(p)
                db.session.flush()
                if stock > 0:
                    db.session.add(StockMovement(
                        product_id=p.id, user_id=current_user.id,
                        movement_type='entrada', quantity=stock,
                        previous_stock=0, new_stock=stock, notes='Importación Excel'
                    ))
                created += 1

        db.session.commit()
        msg = f'Importación completada: {created} creados, {updated} actualizados'
        if skipped:
            msg += f', {skipped} omitidos'
        if errors:
            msg += '. Errores: ' + ' | '.join(errors[:3])
        flash(msg, 'success' if not errors else 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al procesar el archivo: {e}', 'error')

    return redirect(url_for('inventory.products'))


# ─── CATEGORIES ───────────────────────────────────────────────────────────────

@inventory_bp.route('/categorias')
@login_required
def categories():
    cats = Category.query.all()
    return render_template('categories.html', categories=cats)


@inventory_bp.route('/categorias/nueva', methods=['POST'])
@login_required
@editor_required
def new_category():
    name = request.form.get('name', '').strip()
    if not name:
        flash('El nombre de la categoría es obligatorio.', 'error')
        return redirect(url_for('inventory.categories'))
    if Category.query.filter_by(name=name).first():
        flash('Ya existe una categoría con ese nombre.', 'error')
        return redirect(url_for('inventory.categories'))
    cat = Category(
        name=name,
        description=request.form.get('description', ''),
        color=request.form.get('color', '#6366f1')
    )
    db.session.add(cat)
    db.session.commit()
    flash('Categoría creada.', 'success')
    return redirect(url_for('inventory.categories'))


@inventory_bp.route('/categorias/<int:id>/eliminar', methods=['POST'])
@login_required
@admin_required
def delete_category(id):
    cat = Category.query.get_or_404(id)
    if cat.products:
        flash('No se puede eliminar una categoría con productos asignados.', 'error')
        return redirect(url_for('inventory.categories'))
    db.session.delete(cat)
    db.session.commit()
    flash('Categoría eliminada.', 'success')
    return redirect(url_for('inventory.categories'))


# ─── MOVEMENTS ────────────────────────────────────────────────────────────────

@inventory_bp.route('/movimientos')
@login_required
def movements():
    page = request.args.get('page', 1, type=int)
    mvs = StockMovement.query.order_by(
        StockMovement.created_at.desc()
    ).paginate(page=page, per_page=20)
    return render_template('movements.html', movements=mvs)


@inventory_bp.route('/movimientos/exportar')
@login_required
def export_movements_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('inventory.movements'))

    mvs = StockMovement.query.order_by(StockMovement.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Movimientos de Stock'

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='0F766E')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    thin = Border(
        left=Side(style='thin', color='CCFBF1'), right=Side(style='thin', color='CCFBF1'),
        top=Side(style='thin', color='CCFBF1'), bottom=Side(style='thin', color='CCFBF1'),
    )
    TYPE_FILLS = {
        'entrada': PatternFill('solid', fgColor='ECFDF5'),
        'salida': PatternFill('solid', fgColor='FEF2F2'),
        'ajuste': PatternFill('solid', fgColor='FFF7ED'),
    }
    white_f = PatternFill('solid', fgColor='FFFFFF')
    TYPE_LABELS = {'entrada': 'Entrada ↑', 'salida': 'Salida ↓', 'ajuste': 'Ajuste ↔'}

    headers = ['Fecha', 'Producto', 'SKU', 'Tipo', 'Cantidad',
               'Stock Anterior', 'Stock Nuevo', 'Usuario', 'Notas']
    col_widths = [18, 30, 14, 12, 10, 14, 12, 18, 35]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    for i, m in enumerate(mvs, 2):
        fill = TYPE_FILLS.get(m.movement_type, white_f)
        row_data = [
            m.created_at.strftime('%d/%m/%Y %H:%M'),
            m.product.name if m.product else '(eliminado)',
            m.product.sku or '' if m.product else '',
            TYPE_LABELS.get(m.movement_type, m.movement_type),
            m.quantity,
            m.previous_stock if m.previous_stock is not None else '',
            m.new_stock if m.new_stock is not None else '',
            m.user.username if m.user else '',
            m.notes or '',
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin
            cell.fill = fill
            cell.alignment = center if col_idx in [1, 4, 5, 6, 7, 8] else left

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d')
    return send_file(buf, download_name=f'movimientos_stock_{fecha}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
