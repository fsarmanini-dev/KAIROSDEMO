"""Proveedores — CRUD completo + exportar/importar Excel."""
import io
from datetime import datetime

from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, send_file)
from flask_login import login_required

from app.models import db, Proveedor
from app.utils.decorators import admin_required, editor_required

proveedores_bp = Blueprint('proveedores', __name__)


# ─── LISTADO ──────────────────────────────────────────────────────────────────

@proveedores_bp.route('/proveedores')
@login_required
def proveedores():
    q = request.args.get('q', '').strip()
    query = Proveedor.query.filter_by(is_active=True)
    if q:
        query = query.filter(
            db.or_(
                Proveedor.name.ilike(f'%{q}%'),
                Proveedor.contact_name.ilike(f'%{q}%'),
                Proveedor.email.ilike(f'%{q}%'),
                Proveedor.cuit.ilike(f'%{q}%'),
            )
        )
    provs = query.order_by(Proveedor.name).all()
    return render_template('proveedores.html', proveedores=provs, q=q)


# ─── CREAR ────────────────────────────────────────────────────────────────────

@proveedores_bp.route('/proveedores/nuevo', methods=['POST'])
@login_required
@editor_required
def new_proveedor():
    name = request.form.get('name', '').strip()
    if not name:
        flash('El nombre del proveedor es obligatorio.', 'error')
        return redirect(url_for('proveedores.proveedores'))

    p = Proveedor(
        name=name,
        contact_name=request.form.get('contact_name', '').strip(),
        email=request.form.get('email', '').strip(),
        phone=request.form.get('phone', '').strip(),
        address=request.form.get('address', '').strip(),
        cuit=request.form.get('cuit', '').strip(),
        website=request.form.get('website', '').strip(),
        notes=request.form.get('notes', '').strip(),
    )
    db.session.add(p)
    db.session.commit()
    flash(f'Proveedor "{p.name}" creado correctamente.', 'success')
    return redirect(url_for('proveedores.proveedores'))


# ─── EDITAR ───────────────────────────────────────────────────────────────────

@proveedores_bp.route('/proveedores/<int:id>/editar', methods=['POST'])
@login_required
@editor_required
def edit_proveedor(id):
    p = Proveedor.query.get_or_404(id)
    name = request.form.get('name', '').strip()
    if not name:
        flash('El nombre del proveedor es obligatorio.', 'error')
        return redirect(url_for('proveedores.proveedores'))

    p.name = name
    p.contact_name = request.form.get('contact_name', '').strip()
    p.email = request.form.get('email', '').strip()
    p.phone = request.form.get('phone', '').strip()
    p.address = request.form.get('address', '').strip()
    p.cuit = request.form.get('cuit', '').strip()
    p.website = request.form.get('website', '').strip()
    p.notes = request.form.get('notes', '').strip()
    db.session.commit()
    flash(f'Proveedor "{p.name}" actualizado.', 'success')
    return redirect(url_for('proveedores.proveedores'))


# ─── ELIMINAR ─────────────────────────────────────────────────────────────────

@proveedores_bp.route('/proveedores/<int:id>/eliminar', methods=['POST'])
@login_required
@admin_required
def delete_proveedor(id):
    p = Proveedor.query.get_or_404(id)
    p.is_active = False          # soft delete
    db.session.commit()
    flash(f'Proveedor "{p.name}" eliminado.', 'success')
    return redirect(url_for('proveedores.proveedores'))


# ─── EXPORTAR EXCEL ───────────────────────────────────────────────────────────

@proveedores_bp.route('/proveedores/exportar')
@login_required
def export_proveedores_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('proveedores.proveedores'))

    provs = Proveedor.query.filter_by(is_active=True).order_by(Proveedor.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proveedores'

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='0F766E')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin   = Border(
        left=Side(style='thin', color='CCFBF1'),  right=Side(style='thin', color='CCFBF1'),
        top=Side(style='thin', color='CCFBF1'),   bottom=Side(style='thin', color='CCFBF1'),
    )
    alt_fill   = PatternFill('solid', fgColor='F0FDFA')
    white_fill = PatternFill('solid', fgColor='FFFFFF')

    headers    = ['ID', 'Nombre', 'Contacto', 'Email', 'Teléfono', 'CUIT', 'Dirección', 'Web', 'Notas']
    col_widths = [8,    30,       22,         28,      16,         18,     32,           28,     35]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    for i, p in enumerate(provs, 2):
        fill = alt_fill if i % 2 == 0 else white_fill
        row = [p.id, p.name, p.contact_name or '', p.email or '',
               p.phone or '', p.cuit or '', p.address or '',
               p.website or '', p.notes or '']
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border    = thin
            cell.fill      = fill
            cell.alignment = center if col_idx == 1 else left

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d')
    return send_file(buf, download_name=f'proveedores_{fecha}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── DESCARGAR PLANTILLA ──────────────────────────────────────────────────────

@proveedores_bp.route('/proveedores/plantilla')
@login_required
def export_proveedores_template():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('proveedores.proveedores'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proveedores'

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='0F766E')
    center   = Alignment(horizontal='center', vertical='center')
    thin     = Border(
        left=Side(style='thin', color='CCFBF1'),  right=Side(style='thin', color='CCFBF1'),
        top=Side(style='thin', color='CCFBF1'),   bottom=Side(style='thin', color='CCFBF1'),
    )

    headers    = ['Nombre *', 'Contacto', 'Email', 'Teléfono', 'CUIT', 'Dirección', 'Web', 'Notas']
    col_widths = [30,         22,          28,      16,         18,     32,           28,    35]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 24

    # Example row
    example = ['Proveedor Ejemplo S.A.', 'Juan Pérez', 'info@proveedor.com',
               '+54 11 1234-5678', '30-12345678-9', 'Av. Corrientes 1234, CABA',
               'www.proveedor.com', 'Condiciones especiales']
    for col_idx, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(color='94A3B8', italic=True)

    # Instructions sheet
    ws2 = wb.create_sheet('Instrucciones')
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 50
    instr = [
        ('INSTRUCCIONES', ''),
        ('', ''),
        ('Nombre *', 'Obligatorio. Razón social o nombre comercial.'),
        ('Contacto', 'Nombre de la persona de contacto.'),
        ('Email', 'Email de contacto.'),
        ('Teléfono', 'Teléfono principal.'),
        ('CUIT', 'CUIT en formato XX-XXXXXXXX-X.'),
        ('Dirección', 'Dirección completa.'),
        ('Web', 'Sitio web (con o sin https://).'),
        ('Notas', 'Observaciones o condiciones comerciales.'),
        ('', ''),
        ('Nota:', 'Si un proveedor con el mismo nombre ya existe, se actualizarán sus datos.'),
    ]
    for r, (k, v) in enumerate(instr, 1):
        ca = ws2.cell(row=r, column=1, value=k)
        ws2.cell(row=r, column=2, value=v)
        if r == 1:
            ca.font = Font(bold=True, size=12, color='0F766E')
        elif v and r > 2:
            ca.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='plantilla_proveedores.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── IMPORTAR EXCEL ───────────────────────────────────────────────────────────

@proveedores_bp.route('/proveedores/importar', methods=['POST'])
@login_required
@editor_required
def import_proveedores_excel():
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl no está instalado.', 'error')
        return redirect(url_for('proveedores.proveedores'))

    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Por favor subí un archivo Excel (.xlsx).', 'error')
        return redirect(url_for('proveedores.proveedores'))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().rstrip(' *') if cell.value else ''
                   for cell in ws[1]]
        col = {h: i for i, h in enumerate(headers)}

        if 'Nombre' not in col:
            flash('No se encontró la columna "Nombre". Usá la plantilla oficial.', 'error')
            return redirect(url_for('proveedores.proveedores'))

        created = updated = skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            def get(col_name, default=''):
                idx = col.get(col_name)
                if idx is None:
                    return default
                val = row[idx]
                return str(val).strip() if val is not None else default

            name = get('Nombre')
            if not name:
                skipped += 1
                continue

            existing = Proveedor.query.filter(
                Proveedor.name.ilike(name),
                Proveedor.is_active == True
            ).first()

            if existing:
                existing.contact_name = get('Contacto')
                existing.email        = get('Email')
                existing.phone        = get('Teléfono')
                existing.cuit         = get('CUIT')
                existing.address      = get('Dirección')
                existing.website      = get('Web')
                existing.notes        = get('Notas')
                updated += 1
            else:
                db.session.add(Proveedor(
                    name=name,
                    contact_name=get('Contacto'),
                    email=get('Email'),
                    phone=get('Teléfono'),
                    cuit=get('CUIT'),
                    address=get('Dirección'),
                    website=get('Web'),
                    notes=get('Notas'),
                ))
                created += 1

        db.session.commit()
        msg = f'Importación completada: {created} creados, {updated} actualizados'
        if skipped:
            msg += f', {skipped} omitidos'
        flash(msg, 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al procesar el archivo: {e}', 'error')

    return redirect(url_for('proveedores.proveedores'))
